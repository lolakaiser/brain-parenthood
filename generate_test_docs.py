"""
Brain Parenthood – Test Case & Bug Tracking Excel Generator
Covers Sections 4 (Verification) and 5 (Validation) from the SenSym Module 1 spec.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── colour palette ────────────────────────────────────────────────────────────
C_DARK_HEADER  = "1A1A2E"   # dark navy – sheet headers
C_UNIT         = "16213E"   # unit tests
C_INTEGRATION  = "0F3460"   # integration tests
C_SYSTEM       = "533483"   # system tests
C_REGRESSION   = "E94560"   # regression tests
C_UAT          = "1B4332"   # UAT
C_PERF         = "1D3557"   # performance
C_SECURITY     = "7B2D00"   # security
C_USABILITY    = "3D405B"   # usability
C_ROW_ALT      = "F4F6F9"   # alternate row tint
C_WHITE        = "FFFFFF"
C_PASS         = "D4EDDA"
C_FAIL         = "F8D7DA"
C_PENDING      = "FFF3CD"
C_BUG_CRIT     = "F8D7DA"
C_BUG_HIGH     = "FFE5CC"
C_BUG_MED      = "FFF3CD"
C_BUG_LOW      = "D4EDDA"

def hex_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=C_WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def cell_font(size=10, bold=False, color="2C2C2C"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def style_header_row(ws, row, col_widths, bg_color, text_color=C_WHITE):
    for col_idx, _ in enumerate(col_widths, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = hex_fill(bg_color)
        cell.font = header_font(color=text_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

def style_data_row(ws, row, num_cols, alt=False, status_col=None, status_map=None):
    bg = C_ROW_ALT if alt else C_WHITE
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        # status colour override
        if status_col and col_idx == status_col and status_map:
            val = cell.value or ""
            for k, v in status_map.items():
                if k.lower() in str(val).lower():
                    bg = v
                    break
        cell.fill = hex_fill(bg)
        cell.font = cell_font()
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_and_filter(ws, freeze="A3", filter_range=None):
    ws.freeze_panes = freeze
    if filter_range:
        ws.auto_filter.ref = filter_range

# ── DATA ─────────────────────────────────────────────────────────────────────

TC_HEADERS = [
    "Test Case ID", "Type", "Test Objective",
    "Test Steps", "Expected Outcome", "Actual Outcome", "Status", "Notes"
]

VERIFICATION_TESTS = [
    # ── UNIT ──────────────────────────────────────────────────────────────────
    ("TC-U001", "Unit",
     "Verify login with valid email and password",
     "1. Navigate to /login\n2. Enter valid email\n3. Enter correct password\n4. Click Sign In",
     "User is authenticated and redirected to /dashboard",
     "POST /api/users/login returns HTTP 200 with access_token, token_type='bearer', and user object containing email. Verified by automated unit test.", "Pass", "Implemented: __tests__/api/login.test.ts"),

    ("TC-U002", "Unit",
     "Verify login rejects invalid password",
     "1. Navigate to /login\n2. Enter valid email\n3. Enter wrong password\n4. Click Sign In",
     "Error message: 'Invalid credentials' is displayed; user stays on login page",
     "POST /api/users/login returns HTTP 401 with detail='Invalid credentials' for wrong password. Also verified for unknown email. Confirmed by automated unit test.", "Pass", "Implemented: __tests__/api/login.test.ts"),

    ("TC-U003", "Unit",
     "Verify login with username (not email)",
     "1. Navigate to /login\n2. Enter valid username in email field\n3. Enter correct password\n4. Submit",
     "User is authenticated and redirected to /dashboard",
     "When input has no '@', User.findOne is called with {username} instead of {email}. Returns HTTP 200 with valid token. Confirmed by automated unit test.", "Pass", "Implemented: __tests__/api/login.test.ts | Added in commit b1b2fab"),

    ("TC-U004", "Unit",
     "Verify signup creates a new user account",
     "1. Navigate to /signup\n2. Fill in name, email, username, password\n3. Click Create Account",
     "Account created; user receives confirmation; redirected to dashboard",
     "POST /api/users/signup returns HTTP 200 with access_token and user object for a new valid user. User.create called with bcrypt-hashed password. Confirmed by automated unit test.", "Pass", "Implemented: __tests__/api/signup.test.ts"),

    ("TC-U005", "Unit",
     "Verify signup rejects duplicate email",
     "1. Navigate to /signup\n2. Enter email already in DB\n3. Submit",
     "Error: 'Email already in use' is shown; no duplicate document created",
     "Returns HTTP 409 with 'already have an account' message when email exists. Returns HTTP 409 with 'username is already taken' when username is duplicate. User.create is not called. Confirmed by automated unit test.", "Pass", "Implemented: __tests__/api/signup.test.ts"),

    ("TC-U006", "Unit",
     "Verify password strength validation (2-of-3 rule)",
     "1. On signup form, enter passwords with varying strength\n2. Observe live strength indicator",
     "Indicator updates live; submission blocked if < 8 chars or < 2-of-3 criteria met",
     "Returns HTTP 400 for passwords < 8 chars or meeting < 2 of 3 criteria (uppercase/number/special). Accepts passwords with uppercase+number or number+special. 4 variations tested. All pass.", "Pass", "Implemented: __tests__/api/signup.test.ts | 2-of-3: uppercase, number, special char"),

    ("TC-U007", "Unit",
     "Verify forgot-password sends reset email",
     "1. Navigate to /forgot-password\n2. Enter registered email\n3. Submit",
     "API returns success; email dispatched via Resend; reset token stored in DB with 1-hr expiry",
     "For registered email: returns HTTP 200, calls User.findByIdAndUpdate with resetPasswordToken and resetPasswordExpiry fields. For unknown email: returns HTTP 200 silently (prevents account enumeration). Confirmed by automated unit test.", "Pass", "Implemented: __tests__/api/password-reset.test.ts"),

    ("TC-U008", "Unit",
     "Verify password reset with valid token",
     "1. Open reset link from email\n2. Enter new password meeting strength criteria\n3. Submit",
     "Password updated in DB (bcrypt); token invalidated; user redirected to login",
     "POST /api/users/reset-password returns HTTP 200. User.findByIdAndUpdate called with new bcrypt hash and $unset to remove resetPasswordToken and resetPasswordExpiry. Confirmed by automated unit test.", "Pass", "Implemented: __tests__/api/password-reset.test.ts"),

    ("TC-U009", "Unit",
     "Verify reset token is rejected after 1 hour",
     "1. Generate a reset token\n2. Manually set expiry to past\n3. Submit reset form",
     "API returns error: 'Token expired or invalid'",
     "When User.findOne returns null (simulating expired/invalid token), API returns HTTP 400 with 'invalid or has expired' message. No password update occurs. Confirmed by automated unit test.", "Pass", "Implemented: __tests__/api/password-reset.test.ts"),

    ("TC-U010", "Unit",
     "Verify baseline assessment slider saves correct value",
     "1. Go to Module 1 Step 2\n2. Set each slider to a specific value\n3. Proceed to next step",
     "Slider values persist in localStorage and are displayed on Review step",
     "saveBaseline() stores all 7 fields (teamStressLevel, individualStressLevel, productivity, communication, workLifeBalance, teamSize, primaryChallenges) in localStorage. getBaseline() retrieves them intact. Overwrites replace previous values. Timestamp included. 4 tests pass.", "Pass", "Implemented: __tests__/lib/storage.test.ts"),

    ("TC-U011", "Unit",
     "Verify goal-setting form saves answers",
     "1. Go to Module 1 Step 3\n2. Fill all 6 goal fields\n3. Click Next",
     "Answers saved to brainParenthood_goals in localStorage",
     "saveGoals() persists all 6 fields (stressReduction, productivityGoal, communicationGoal, personalGoal, teamGoal, successMetrics) to localStorage key 'brainParenthood_goals'. getGoals() retrieves them correctly. completedAt timestamp included. 3 tests pass.", "Pass", "Implemented: __tests__/lib/storage.test.ts"),

    ("TC-U012", "Unit",
     "Verify module progress marks as complete on final step",
     "1. Complete all 5 steps of Module 1\n2. Click Submit on Review step",
     "Module 1 added to completedModules[]; progress bar updates to reflect completion",
     "completeModule(1) adds 1 to completedModules[] and sets currentModule=2. Completing modules in any order keeps array sorted. Duplicate calls do not create duplicates. isModuleCompleted() returns correct boolean. getProgress() returns defaults {[],1} when empty. 8 tests pass.", "Pass", "Implemented: __tests__/lib/storage.test.ts"),

    # ── INTEGRATION ───────────────────────────────────────────────────────────
    ("TC-I001", "Integration",
     "Login API integrates with MongoDB for credential verification",
     "1. POST to /api/users/login with valid credentials\n2. Inspect MongoDB query and response",
     "bcryptjs compare runs against stored hash; base64 token returned on match",
     "", "Pending", ""),

    ("TC-I002", "Integration",
     "Signup API creates user document in MongoDB",
     "1. POST to /api/users/signup\n2. Verify document in MongoDB users collection",
     "Document includes hashed password (bcrypt), email, username, name; moduleProgress initialised",
     "", "Pending", ""),

    ("TC-I003", "Integration",
     "Progress API syncs localStorage data to MongoDB",
     "1. Complete a module on client\n2. POST to /api/progress\n3. Verify DB document updated",
     "moduleProgress.completedModules and currentModule updated in user document",
     "", "Pending", ""),

    ("TC-I004", "Integration",
     "Answers API persists assessment/goal data in MongoDB",
     "1. POST to /api/answers with moduleId=1, step='assessment', answers={...}\n2. Verify DB",
     "moduleAnswers['1'].assessment stored; savedAt timestamp written",
     "", "Pending", ""),

    ("TC-I005", "Integration",
     "Forgot-password API integrates with Resend email service",
     "1. POST to /api/users/forgot-password\n2. Monitor Resend API call and response",
     "Resend API called with correct to/from/subject/html; 200 returned regardless of email existence",
     "", "Pending", "Account enumeration prevention"),

    ("TC-I006", "Integration",
     "AuthContext validates token via /api/users/me on app load",
     "1. Set valid authToken in localStorage\n2. Reload app\n3. Observe AuthContext behaviour",
     "GET /api/users/me called; user state hydrated; progress loaded from MongoDB",
     "", "Pending", ""),

    ("TC-I007", "Integration",
     "Module completion triggers progress sync to DB",
     "1. Submit final Review step\n2. Observe localStorage update\n3. Verify POST to /api/progress",
     "completedModules updated both locally and in MongoDB within 2 seconds",
     "", "Pending", "Fire-and-forget sync"),

    # ── SYSTEM ────────────────────────────────────────────────────────────────
    ("TC-S001", "System",
     "End-to-end: new user registers and reaches dashboard",
     "1. Open app\n2. Click Sign Up\n3. Complete signup form\n4. Verify dashboard loads with empty progress",
     "User sees dashboard with 0% progress; Module 1 unlocked; Modules 2-12 locked",
     "", "Pending", ""),

    ("TC-S002", "System",
     "End-to-end: complete Module 1 through all 5 steps",
     "1. Login\n2. Go to Module 1\n3. Complete Overview → Assessment → Goals → Review → Complete",
     "Module 1 marked complete; badge turns green; Module 2 unlocked; dashboard progress updates",
     "", "Pending", ""),

    ("TC-S003", "System",
     "Module lock system prevents skipping ahead",
     "1. Login as new user\n2. Manually navigate to /module/3\n3. Observe behaviour",
     "Non-admin user redirected or shown locked state; cannot submit answers for locked module",
     "", "Pending", "Admin bypass should still work"),

    ("TC-S004", "System",
     "Password reset end-to-end flow works correctly",
     "1. Click Forgot Password\n2. Enter email\n3. Open email link\n4. Enter new password\n5. Login",
     "Login succeeds with new password; old password rejected",
     "", "Pending", ""),

    ("TC-S005", "System",
     "Dashboard stats reflect actual module completion",
     "1. Complete modules 1, 2, 3\n2. Go to dashboard",
     "Progress bar shows 25% (3/12); completed count = 3; currentModule = 4",
     "", "Pending", ""),

    ("TC-S006", "System",
     "Admin user can access any module regardless of progress",
     "1. Login as admin (isAdmin=true)\n2. Navigate to /module/12 without completing earlier modules",
     "Admin accesses Module 12 without redirect; all modules show as unlocked",
     "", "Pending", ""),

    # ── REGRESSION ────────────────────────────────────────────────────────────
    ("TC-R001", "Regression",
     "Login still works after password reset",
     "1. Reset password via forgot-password flow\n2. Login with new password",
     "Login succeeds; old password no longer works",
     "", "Pending", "Regression for commit 495e808"),

    ("TC-R002", "Regression",
     "Completed modules remain complete after logout/login",
     "1. Complete Module 1\n2. Logout\n3. Login again\n4. Check dashboard",
     "Module 1 still shows as complete; progress loaded from MongoDB",
     "", "Pending", ""),

    ("TC-R003", "Regression",
     "Clickable step navigation dots work after editing a step",
     "1. Go to Module 1\n2. Advance to Step 4 (Review)\n3. Click a dot to go back to Step 2\n4. Edit an answer\n5. Navigate forward using dots",
     "Dots navigate correctly; edited answer persists through navigation",
     "", "Pending", "Regression for commit 9d1190e"),

    ("TC-R004", "Regression",
     "Edit buttons on review page jump to correct question",
     "1. Complete all module steps to Review\n2. Click Edit button next to Question 3",
     "App navigates directly to Step 2 with focus on Question 3 input",
     "", "Pending", "Regression for commit 424a0c3"),

    ("TC-R005", "Regression",
     "Username login does not break email login",
     "1. Login with email\n2. Logout\n3. Login with username",
     "Both login methods succeed; same user account returned",
     "", "Pending", "Regression for commit b1b2fab"),
]

VALIDATION_TESTS = [
    # ── UAT ───────────────────────────────────────────────────────────────────
    ("TC-UAT001", "UAT",
     "New user can complete full onboarding without assistance",
     "1. Recruit 5 first-time users\n2. Ask them to register and complete Module 1\n3. Observe without help",
     "All testers complete registration and Module 1 within 15 minutes; no blocking confusion",
     "", "Pending", "Social/environmental impact: ensure accessible language"),

    ("TC-UAT002", "UAT",
     "User can intuitively navigate between all 12 modules",
     "1. Login as user with some progress\n2. Ask to jump to any module\n3. Observe navigation",
     "User locates correct module via /modules page or dashboard within 30 seconds",
     "", "Pending", ""),

    ("TC-UAT003", "UAT",
     "User can review and edit answers before final submission",
     "1. Complete Steps 1-3 of a module\n2. On Review step, click Edit on any answer\n3. Change value\n4. Re-review",
     "Edited value shown in review; submission reflects updated answer",
     "", "Pending", ""),

    ("TC-UAT004", "UAT",
     "Forgot-password email is received and usable by real users",
     "1. Use real email address\n2. Trigger forgot-password flow\n3. Check inbox",
     "Email arrives within 2 minutes; link works; password successfully reset",
     "", "Pending", "Test with gmail, outlook, yahoo"),

    ("TC-UAT005", "UAT",
     "Dashboard provides meaningful, motivating progress feedback",
     "1. Complete 3 modules\n2. Open dashboard\n3. Survey users on clarity of progress display",
     "≥80% of test users rate progress display as 'clear' or 'very clear'",
     "", "Pending", "Responsible innovation: avoid anxiety-inducing metrics"),

    ("TC-UAT006", "UAT",
     "Buy Me a Coffee donation flow works end-to-end",
     "1. Click coffee cup icon in nav\n2. Select amount\n3. Enter Square test card details\n4. Submit",
     "Payment processed; confirmation shown; no errors in console",
     "", "Pending", "Use Square sandbox environment"),

    ("TC-UAT007", "UAT",
     "Logout clears session and prevents re-entry without login",
     "1. Login\n2. Click logout\n3. Use browser back button or navigate to /dashboard",
     "User cannot access protected pages; redirected to /login",
     "", "Pending", ""),

    # ── PERFORMANCE ───────────────────────────────────────────────────────────
    ("TC-P001", "Performance",
     "Login API responds within 2 seconds under normal load",
     "1. Use tool (e.g. Postman/k6)\n2. Send 50 concurrent POST /api/users/login requests",
     "95th percentile response time < 2000ms; no 5xx errors",
     "", "Pending", "bcryptjs cost factor may slow under high concurrency"),

    ("TC-P002", "Performance",
     "Dashboard page loads within 3 seconds",
     "1. Open Chrome DevTools → Lighthouse\n2. Run performance audit on /dashboard",
     "Performance score ≥ 80; LCP < 3s; TTI < 3s",
     "", "Pending", ""),

    ("TC-P003", "Performance",
     "Module page with saved answers loads within 2 seconds",
     "1. Complete several module steps\n2. Navigate back to a completed module",
     "Page fully interactive in < 2 seconds; no layout shift on answer hydration",
     "", "Pending", ""),

    ("TC-P004", "Performance",
     "System handles 100 concurrent users without degradation",
     "1. Simulate 100 concurrent users with k6 or similar\n2. Run mixed login + progress-save scenario",
     "Error rate < 1%; avg response time < 3s; MongoDB connection pool not exhausted",
     "", "Pending", ""),

    ("TC-P005", "Performance",
     "MongoDB queries complete within 500ms",
     "1. Enable Mongoose query logging\n2. Trigger user fetch, progress save, answers save",
     "All queries complete in < 500ms; indexes on email and username are in place",
     "", "Pending", ""),

    # ── SECURITY ──────────────────────────────────────────────────────────────
    ("TC-SEC001", "Security",
     "Login form is protected against NoSQL injection",
     "1. In login email field enter: {\"$gt\": \"\"}\n2. Submit",
     "Request rejected or sanitised; no unauthorised authentication",
     "", "Pending", "Mongoose sanitisation by default; verify explicitly"),

    ("TC-SEC002", "Security",
     "Text inputs are protected against XSS",
     "1. Enter <script>alert('xss')</script> in Primary Challenges textarea\n2. Submit and review",
     "Script not executed; value stored/displayed as escaped text",
     "", "Pending", "React escapes by default; verify no dangerouslySetInnerHTML"),

    ("TC-SEC003", "Security",
     "Manipulated auth token is rejected",
     "1. Decode base64 token\n2. Change userId to another user's ID\n3. Re-encode and call /api/users/me",
     "API returns 401; user cannot access another user's data",
     "", "Pending", "Current token is base64 only – priority security finding"),

    ("TC-SEC004", "Security",
     "Unauthenticated requests to protected API routes return 401",
     "1. Call GET /api/progress without Authorization header",
     "API returns 401 Unauthorized; no data leaked",
     "", "Pending", ""),

    ("TC-SEC005", "Security",
     "Passwords are stored as bcrypt hashes in MongoDB",
     "1. Create an account\n2. Inspect user document in MongoDB",
     "Password field contains a $2b$ bcrypt hash; plaintext never stored",
     "", "Pending", "Salt rounds = 12"),

    ("TC-SEC006", "Security",
     "Password reset token expires after 1 hour",
     "1. Request password reset\n2. Wait 61 minutes\n3. Use the reset link",
     "API returns error indicating token expired; token deleted from DB",
     "", "Pending", ""),

    ("TC-SEC007", "Security",
     "Application enforces HTTPS in production",
     "1. Access the production URL via HTTP\n2. Observe redirect",
     "Browser redirected to HTTPS; HSTS header present",
     "", "Pending", "Verify on deployment environment"),

    ("TC-SEC008", "Security",
     "Login endpoint has rate limiting to prevent brute force",
     "1. Send 20+ failed login attempts in quick succession",
     "After threshold, endpoint returns 429 Too Many Requests",
     "", "Pending", "Currently not implemented – open risk item"),

    # ── USABILITY ─────────────────────────────────────────────────────────────
    ("TC-USE001", "Usability",
     "Password strength indicator is visible and understandable",
     "1. Navigate to /signup\n2. Begin typing a password",
     "Strength bar updates in real time; colour and label clearly communicate strength level",
     "", "Pending", ""),

    ("TC-USE002", "Usability",
     "Error messages are clear and actionable",
     "1. Trigger validation errors on login, signup, and reset-password forms",
     "Each error message explains what went wrong and how to fix it (not just 'Error')",
     "", "Pending", ""),

    ("TC-USE003", "Usability",
     "All pages are responsive on mobile (375px viewport)",
     "1. Open DevTools\n2. Set viewport to 375px wide\n3. Visit all key pages",
     "No horizontal scroll; text legible; buttons tappable (≥44px); no overlapping elements",
     "", "Pending", ""),

    ("TC-USE004", "Usability",
     "Step indicator shows current step clearly",
     "1. Navigate through Module 1 steps",
     "Active step dot/indicator visually distinct; step number/name displayed",
     "", "Pending", ""),

    ("TC-USE005", "Usability",
     "Progress bar accurately reflects module completion percentage",
     "1. Complete various numbers of modules\n2. Check dashboard progress bar",
     "Bar shows correct percentage (completed/12 × 100); number label matches",
     "", "Pending", ""),

    ("TC-USE006", "Usability",
     "Key actions are accessible via keyboard navigation",
     "1. Tab through login, signup, and module forms\n2. Submit using Enter key",
     "All interactive elements reachable via Tab; forms submittable without mouse",
     "", "Pending", "WCAG 2.1 AA compliance"),

    ("TC-USE007", "Usability",
     "Form validation feedback appears immediately on blur",
     "1. Click into an input, then click away without filling it",
     "Inline error appears within 300ms; does not wait for full form submission",
     "", "Pending", ""),
]

BUG_HEADERS = [
    "Bug ID", "Date Reported", "Reporter", "Feature / Module",
    "Bug Title", "Description", "Steps to Reproduce",
    "Expected Behaviour", "Actual Behaviour",
    "Severity", "Priority", "Status",
    "Assigned To", "Date Fixed", "Resolution Notes"
]

BUG_DATA = [
    ("BUG-001", "2025-01-15", "QA Team", "Authentication",
     "Login rejects valid username containing uppercase letters",
     "Users whose usernames contain uppercase characters (e.g. 'JohnDoe') cannot login with their username even though signup allows mixed-case usernames.",
     "1. Sign up with username 'JohnDoe'\n2. Logout\n3. Login with username 'JohnDoe'",
     "User authenticated successfully",
     "Error: 'Invalid credentials'",
     "High", "P1", "Fixed",
     "Dev Team", "2025-01-16", "Added .toLowerCase() normalisation in login handler"),

    ("BUG-002", "2025-01-20", "User Report", "Module Progress",
     "Completed module badge not displayed after page refresh",
     "After completing a module and refreshing the /modules page, the green 'Completed' badge reverts to the default state until the user navigates away and back.",
     "1. Complete Module 1\n2. Navigate to /modules\n3. Hard refresh (Ctrl+R)",
     "Module 1 shows green Completed badge",
     "Module 1 shows default state; badge missing",
     "Medium", "P2", "Fixed",
     "Dev Team", "2025-01-22", "Progress hydration from MongoDB now runs before first paint on /modules"),

    ("BUG-003", "2025-02-01", "QA Team", "Module 1 – Review Step",
     "Edit button on Review step scrolls to wrong question",
     "Clicking the Edit button next to Question 5 (Work-Life Balance) navigates to Question 4 (Communication) instead.",
     "1. Complete Module 1 Steps 1-3\n2. On Review (Step 4), click Edit next to Work-Life Balance",
     "Navigates to Work-Life Balance input (Question 5)",
     "Navigates to Communication input (Question 4)",
     "Medium", "P2", "Fixed",
     "Dev Team", "2025-02-05", "Off-by-one error in question index array – corrected in commit 424a0c3"),

    ("BUG-004", "2025-02-10", "Security Audit", "Authentication",
     "Auth token uses base64 encoding with no signature – susceptible to user impersonation",
     "The authentication token is a plain base64-encoded string containing userId:email:timestamp with no HMAC or cryptographic signature. An attacker who intercepts or guesses a token can modify the userId to access another user's data.",
     "1. Obtain any valid auth token\n2. Decode from base64\n3. Replace userId with another user's MongoDB _id\n4. Re-encode and send to GET /api/users/me",
     "401 Unauthorized",
     "200 OK with victim user's profile data",
     "Critical", "P0", "Open",
     "Dev Team", "", "Recommend migrating to signed JWT (jsonwebtoken) with HS256 or RS256"),

    ("BUG-005", "2025-02-14", "QA Team", "Signup",
     "Duplicate username check is case-sensitive – allows 'Admin' and 'admin' as separate accounts",
     "MongoDB unique index on username is case-sensitive by default. A user can register with 'Admin' after 'admin' already exists, creating two accounts that visually look identical.",
     "1. Register with username 'admin'\n2. Register again with username 'Admin'",
     "Error: 'Username already taken'",
     "Second account created successfully",
     "Medium", "P2", "Open",
     "Dev Team", "", "Add collation {locale:'en', strength:2} to username index or normalise to lowercase on save"),

    ("BUG-006", "2025-02-18", "User Report", "Dashboard",
     "Progress percentage shows NaN when user has no completed modules",
     "First-time users who have not completed any module see 'NaN%' on the dashboard progress card instead of '0%'.",
     "1. Create a new account\n2. Navigate to /dashboard",
     "Progress shows '0%'",
     "Progress shows 'NaN%'",
     "Low", "P3", "Fixed",
     "Dev Team", "2025-02-19", "Added null coalesce: (completedModules?.length ?? 0) / 12 * 100"),

    ("BUG-007", "2025-02-25", "QA Team", "Password Reset",
     "Reset password page accepts whitespace-only passwords",
     "The password reset form does not trim whitespace; a password of '        ' (8 spaces) passes the length check and is accepted.",
     "1. Request password reset\n2. Open reset link\n3. Enter '        ' (8 spaces) as password\n4. Submit",
     "Validation error: password does not meet complexity requirements",
     "Password accepted and updated in DB",
     "High", "P1", "Open",
     "Dev Team", "", "Add .trim() before length/complexity validation in reset-password API handler"),

    ("BUG-008", "2025-03-01", "Performance Test", "API – Login",
     "Login endpoint blocks under high concurrency due to bcrypt cost factor 12",
     "Under load of 100 concurrent login requests, the Node.js event loop is blocked by synchronous bcrypt.compare operations, causing latency spikes above 10 seconds.",
     "1. Run k6 load test with 100 VUs hitting POST /api/users/login\n2. Observe response times",
     "95th percentile < 2000ms",
     "95th percentile ~12,000ms; intermittent 504 errors",
     "High", "P1", "Open",
     "Dev Team", "", "Use bcrypt.compare async (already async in bcryptjs); consider reducing cost to 10 or adding a request queue"),

    ("BUG-009", "2025-03-05", "QA Team", "Navigation",
     "Step indicator dots are not keyboard accessible",
     "The clickable step navigation dots added in commit 9d1190e are implemented as <div> elements with onClick handlers. They cannot be reached or activated via keyboard Tab/Enter.",
     "1. Open Module 1\n2. Tab through the page\n3. Attempt to navigate using step dots",
     "Dots focusable and activatable via keyboard",
     "Dots are not in tab order; keyboard users cannot use step navigation",
     "Medium", "P2", "Open",
     "Dev Team", "", "Change <div> to <button> or add tabIndex={0} and onKeyDown handler for Enter/Space"),

    ("BUG-010", "2025-03-10", "User Report", "Module 1",
     "Slider values reset to default when using browser back button",
     "After setting slider values in the Baseline Assessment (Step 2) and clicking Next, pressing the browser back button resets all sliders to their default midpoint value (5) instead of restoring entered values.",
     "1. Go to Module 1 Step 2\n2. Set Team Stress Level slider to 8\n3. Click Next\n4. Click browser back button",
     "Sliders restored to previously entered values",
     "All sliders reset to default value of 5",
     "Medium", "P2", "Open",
     "Dev Team", "", "Read initial slider values from localStorage on component mount"),
]

# ── WORKBOOK BUILDER ──────────────────────────────────────────────────────────

def build_title_sheet(wb):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "Brain Parenthood – Test Case & Bug Tracking Documentation"
    title.font = Font(name="Calibri", size=20, bold=True, color=C_WHITE)
    title.fill = hex_fill(C_DARK_HEADER)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48

    meta = [
        ("Project",       "Brain Parenthood – 12-Week Resilience Toolkit"),
        ("Version",       "1.0"),
        ("Date",          "2026-03-22"),
        ("Prepared by",   "SenSym Module 1 QA"),
        ("Methodology",   "Sections 4 (Verification) & 5 (Validation) per SenSym spec"),
        ("Scope",         "Authentication · Module Progression · Data Sync · UI/UX · Security"),
        ("Sheets",        "Cover | Verification Tests | Validation Tests | Bug Tracker"),
    ]
    for r, (k, v) in enumerate(meta, start=3):
        ws.cell(row=r, column=2, value=k).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=r, column=3, value=v).font = Font(name="Calibri", size=11)
        ws.row_dimensions[r].height = 20

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 60


def build_test_sheet(wb, sheet_name, tests, section_header, type_color_map):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(TC_HEADERS))}1")
    title = ws["A1"]
    title.value = section_header
    title.font = Font(name="Calibri", size=14, bold=True, color=C_WHITE)
    title.fill = hex_fill(C_DARK_HEADER)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # Column headers
    for col, h in enumerate(TC_HEADERS, start=1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, TC_HEADERS, C_DARK_HEADER)
    ws.row_dimensions[2].height = 28

    status_map = {"pass": C_PASS, "fail": C_FAIL, "pending": C_PENDING}

    last_type = None
    row = 3  # manual row counter so separator rows don't collide with data rows
    for tc in tests:
        tc_type = tc[1]
        # insert a coloured separator row each time the type changes
        if tc_type != last_type:
            color = type_color_map.get(tc_type, C_DARK_HEADER)
            ws.merge_cells(f"A{row}:{get_column_letter(len(TC_HEADERS))}{row}")
            sep = ws.cell(row=row, column=1,
                          value=f"  ▶  {tc_type.upper()} TEST CASES")
            sep.font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
            sep.fill = hex_fill(color)
            sep.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 22
            last_type = tc_type
            row += 1

        for col, val in enumerate(tc, start=1):
            ws.cell(row=row, column=col, value=val)
        alt = (row % 2 == 0)
        style_data_row(ws, row, len(TC_HEADERS), alt, status_col=7, status_map=status_map)
        ws.row_dimensions[row].height = 80
        row += 1

    col_widths = [14, 14, 32, 42, 32, 20, 12, 20]
    set_col_widths(ws, col_widths)
    freeze_and_filter(ws, "A3", f"A2:{get_column_letter(len(TC_HEADERS))}2")


def build_bug_sheet(wb):
    ws = wb.create_sheet("Bug Tracker")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(BUG_HEADERS))}1")
    title = ws["A1"]
    title.value = "Bug Tracking Log – Brain Parenthood"
    title.font = Font(name="Calibri", size=14, bold=True, color=C_WHITE)
    title.fill = hex_fill(C_DARK_HEADER)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    for col, h in enumerate(BUG_HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, BUG_HEADERS, C_DARK_HEADER)
    ws.row_dimensions[2].height = 28

    severity_map = {
        "critical": C_BUG_CRIT,
        "high": C_BUG_HIGH,
        "medium": C_BUG_MED,
        "low": C_BUG_LOW,
    }
    status_map = {
        "fixed": C_PASS,
        "closed": C_PASS,
        "open": C_FAIL,
        "in progress": C_PENDING,
    }

    for row_idx, bug in enumerate(BUG_DATA, start=3):
        for col, val in enumerate(bug, start=1):
            ws.cell(row=row_idx, column=col, value=val)

        # severity colouring on col 10
        sev = str(bug[9]).lower()
        sev_bg = severity_map.get(sev, C_WHITE)
        # status colouring on col 12
        stat = str(bug[11]).lower()
        stat_bg = status_map.get(stat, C_WHITE)

        alt = (row_idx % 2 == 0)
        base_bg = C_ROW_ALT if alt else C_WHITE
        for col in range(1, len(BUG_HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col)
            bg = sev_bg if col == 10 else (stat_bg if col == 12 else base_bg)
            cell.fill = hex_fill(bg)
            cell.font = cell_font()
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 100

    bug_col_widths = [10, 13, 13, 18, 30, 38, 38, 28, 28, 11, 10, 13, 14, 13, 35]
    set_col_widths(ws, bug_col_widths)
    freeze_and_filter(ws, "A3", f"A2:{get_column_letter(len(BUG_HEADERS))}2")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_title_sheet(wb)

    # Verification sheet
    ver_color_map = {
        "Unit":        C_UNIT,
        "Integration": C_INTEGRATION,
        "System":      C_SYSTEM,
        "Regression":  C_REGRESSION,
    }
    build_test_sheet(
        wb,
        "Verification Tests",
        VERIFICATION_TESTS,
        "Section 4 – Verification Test Cases (Unit · Integration · System · Regression)",
        ver_color_map,
    )

    # Validation sheet
    val_color_map = {
        "UAT":         C_UAT,
        "Performance": C_PERF,
        "Security":    C_SECURITY,
        "Usability":   C_USABILITY,
    }
    build_test_sheet(
        wb,
        "Validation Tests",
        VALIDATION_TESTS,
        "Section 5 – Validation Test Cases (UAT · Performance · Security · Usability)",
        val_color_map,
    )

    build_bug_sheet(wb)

    out_path = r"C:\Users\lolak\OneDrive\Desktop\SenSym Module 1\BrainParenthood_TestCases_BugLog.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
